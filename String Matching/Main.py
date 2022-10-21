# -*- coding: utf-8 -*-
"""
Created on Fri Nov 26 22:19:24 2021

@author: udyan.sachdev
"""


import pandas as pd
import os
import sys
from openpyxl import load_workbook
from fuzzywuzzy import fuzz                      
from nltk.corpus import stopwords
from datetime import date
from tabulate import tabulate             #pip install tabulate

# pip install fuzzywuzzy 
#pip install python-Levenshtein
   
cwd = os.getcwd()
print('')
print("Current working directory: {0}".format(cwd))

inp = input('Do you want to continue at this directory- Y or N : ')
files_xls=[]
if (inp == 'Y' or inp == 'y' or inp =='Yes' or inp =='yes'):
    print(' ')
    file = os.listdir(cwd)
    files_xls1 = [f for f in file if (f[-3:] == 'csv') or (f[-3:] == 'xls')]
    files_xls2 = [f for f in file if f[-4:] == 'xlsx']

    files_xls = files_xls1 + files_xls2 
    
    if len(files_xls) == 0:
        print('No excel Found! Change the directory and try again')
        
    else:
        try:
            print('Here is the list of available excels : ')

        
            print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(files_xls))]))
            print(' ')
        
            num = int(input("Enter the serial number corresponding the required excel: "))
            
            if files_xls[num][-3:] == 'csv':
                full_path = cwd + '\\' + files_xls[num]
        
                data = pd.read_csv(full_path)
                print('All the columns in this worksheet are : ')
                print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data.columns))]))
                print('')
                col_num= int(input('Select the serial number of the single column to be used for string matching -'))
                col_name = data.columns[col_num]   
                
            elif files_xls[num][-3:] == 'xls':
                full_path = cwd + '\\' + files_xls[num]
        
                data = pd.read_excel(full_path)
                print('All the columns in this worksheet are : ')
                print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data.columns))]))
                print('')
                col_num= int(input('Select the serial number of the single column to be used for string matching -'))
                col_name = data.columns[col_num]
                
            else:
                full_path = cwd + '\\' + files_xls[num]
                
                wb = load_workbook(filename = files_xls[num])
                print(' ')
                print('Here is the list of available worksheet : ')
                print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(wb.sheetnames))]))
                print(' ')
                
                sn = int(input("Enter the serial number corresponding the required Sheet name: "))
                
                data = pd.read_excel(full_path,sheet_name = wb.sheetnames[sn])
                print('All the columns in this worksheet are : ')
                print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data.columns))]))
                print('')
                col_num= int(input('Select the serial number of the single column to be used for string matching -'))
                col_name = data.columns[col_num]
 
                
        except IndexError:
            print('Input the number in the range')
            sys.exit()
                
    #Data Cleaning:
        
    #The columns should be in string 
    data[col_name] = data[col_name].astype(str)
        
    #Filtering only unique values and Removing nan values
    data = data.drop_duplicates(subset = [col_name])
    
    # replace symbols: \n “ ‘ / ( ) { } | @ , ; # with space
    data[col_name] = data[col_name].str.replace('\n “ ‘ / ( ) { } | @ , ; #','',regex=True)
    
    
    # remove leading and trailing spaces
    data[col_name] = data[col_name].str.strip()
    
    # remove stop words
    stop = stopwords.words('english') 
    data[col_name] = data[col_name].apply(lambda x: ' '.join([word for word in x.split() if word not in (stop)]))
        
    inp = int(input('1 - Compare with a master list ' +'\n' + '2 - Compare column of an another dataframe ' + '\n' + '3 - Compare the column with itself '+ '\n' 'Enter the required serial number -'))
    
    
    print(tabulate([['1', 'ratio','Two input strings of almost same length'], ['2','partial_ratio','Attempts to account for partial string matches better'],
                    ['3','token_sort_ratio','Attempts to account for similar strings out of order'],['4','token_set_ratio','Attempts to rule out differences in the strings'],
                    ['5','partial_token_sort_ratio','Same algorithm as in token_sort_ratio, but instead of applying ratio after sorting the tokens, uses partial_ratio'],['6','partial_token_set_ratio','Same algorithm as in token_set_ratio, but instead of applying ratio to the sets of tokens, uses partial_ratio'],
                    ['7','WRatio','An attempt to weight (the name stands for "Weighted Ratio") results from different algorithms to calculate the score.']], headers=['Serial No.', 'Fuzzywuzzy Algorithms' , 'When to use'], tablefmt='youtrack'))
    print('')
    print('Refer to this link for more info - https://chairnerd.seatgeek.com/fuzzywuzzy-fuzzy-string-matching-in-python/')
    print('')
    
    algo = int(input('Select the serial number of the required algorithm to be used - '))
    
    if (algo==1):
        method = fuzz.ratio
    elif(algo==2):
        method = fuzz.partial_ratio
    elif(algo==3):
        method = fuzz.token_sort_ratio
    elif(algo==4):
        method = fuzz.token_set_ratio
    elif(algo==5):
        method = fuzz.partial_token_sort_ratio
    elif(algo==6):
        method = fuzz.partial_token_set_ratio
    elif(algo==7):
        method = fuzz.WRatio
    else:
        print('Enter the algo in the table')
        sys.exit()
    
    if (inp==1):     
           
        master_list = input('Enter the master list seperated by commas - ')
        master_list = list(master_list.split(","))
        master_list = [x.strip(' ') for x in master_list]
        score = int(input('Enter the minimum cutoff score required on a scale of 0 - 100 (100 being max ,Ideal cutoff score is 80) - '))
        
        #Fuzzywuzzy algorithm
         
        compare = pd.MultiIndex.from_product([master_list,data[col_name]]).to_series()
        
        def metrics(tup):
            return pd.Series([method(*tup)],['Score'])
            
        compare = compare.apply(metrics)
        compare_scr=compare[(compare['Score'] > score )]
        compare_scr.to_csv('Str_Matching_' + method.__name__ + '_on_' + col_name + '_' + date.today().strftime("%d%b%y") + '.csv')
        print('Excel has been created with the results!!')
    
            
    elif(inp==2):
        
        print('')
        print('Selecting the second dataframe')
        cwd = os.getcwd()
        print('')
        print("Current working directory: {0}".format(cwd))
        print(' ')
        
        file = os.listdir(cwd)
        files_xls1 = [f for f in file if (f[-3:] == 'csv') or (f[-3:] == 'xls')]
        files_xls2 = [f for f in file if f[-4:] == 'xlsx']
    
        files_xls = files_xls1 + files_xls2 
        
        if len(files_xls) == 0:
            print('No excel Found! Change the directory and try again')
            
        else:
            try:
                print('Here is the list of available excels : ')
    
            
                print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(files_xls))]))
                print(' ')
            
                num = int(input("Enter the serial number corresponding the required excel: "))
                
                if files_xls[num][-3:] == 'csv':
                    full_path = cwd + '\\' + files_xls[num]
            
                    data1 = pd.read_csv(full_path)
                    print('All the columns in this worksheet are : ')
                    print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data1.columns))]))
                    print('')
                    col_num = int(input('Select the serial number of the single column to be used for string matching -'))
                    col_name1 = data1.columns[col_num]   
                    
                elif files_xls[num][-3:] == 'xls':
                    full_path = cwd + '\\' + files_xls[num]
            
                    data1 = pd.read_excel(full_path)
                    print('All the columns in this worksheet are : ')
                    print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data1.columns))]))
                    print('')
                    col_num= int(input('Select the serial number of the single column to be used for string matching -'))
                    col_name1 = data1.columns[col_num]
                    
                else:
                    full_path = cwd + '\\' + files_xls[num]
                    
                    wb = load_workbook(filename = files_xls[num])
                    print(' ')
                    print('Here is the list of available worksheet : ')
                    print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(wb.sheetnames))]))
                    print(' ')
                    
                    sn = int(input("Enter the serial number corresponding the required Sheet name: "))
                    
                    data1 = pd.read_excel(full_path,sheet_name = wb.sheetnames[sn])
                    print('All the columns in this worksheet are : ')
                    print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data1.columns))]))
                    print('')
                    col_num= int(input('Select the serial number of the single column to be used for string matching -'))
                    col_name1 = data1.columns[col_num]
         
                        
            except IndexError:
                print('Input the number in the range')
                sys.exit()
    
        #Data Cleaning:
            
        #The columns should be in string 
        data1[col_name1] = data1[col_name1].astype(str)
            
        #Filtering only unique values and Removing nan values
        data1 = data1.drop_duplicates(subset = [col_name1])
        
        # replace symbols: \n “ ‘ / ( ) { } | @ , ; # with space
        data1[col_name1] = data1[col_name1].str.replace('\n “ ‘ / ( ) { } | @ , ; #','',regex=True)
    
    
        # remove leading and trailing spaces
        data1[col_name1] = data1[col_name1].str.strip()
    
        # remove stop words
        stop = stopwords.words('english') 
        data1[col_name1] = data1[col_name1].apply(lambda x: ' '.join([word for word in x.split() if word not in (stop)]))
    
        score = int(input('Enter the minimum cutoff score required on a scale of 0 - 100 (100 being max ,Ideal cutoff score is 80) - '))
    
        #Fuzzywuzzy algorithm
        matched_vendors = []
    
        for row in data.index:
            vendor_name = data._get_value(row,col_name)
            for columns in data1.index:
                regulated_vendor_name=data1._get_value(columns,col_name1)
                matched_token=method(vendor_name,regulated_vendor_name)
                if matched_token > score:
                    matched_vendors.append([vendor_name,regulated_vendor_name,matched_token])
        
       
        pd.DataFrame(matched_vendors,columns=[col_name,col_name1,'Score']).to_csv('Str_Matching_'+ method.__name__ + '_on_' + col_name + '_' + date.today().strftime("%d%b%y") + '.csv', index=False)
        print('Excel has been created with the results!!')
    
    
    elif(inp==3):
        score = int(input('Enter the minimum cutoff score required on a scale of 0 - 100 (100 being max, Ideal cutoff score is 80) - '))
    
        #Fuzzywuzzy algorithm
        matched_vendors = []
    
        for row in data.index:
            vendor_name = data._get_value(row,col_name)
            for columns in data.index:
                regulated_vendor_name=data._get_value(columns,col_name)
                matched_token=method(vendor_name,regulated_vendor_name)
                if matched_token > score:
                    matched_vendors.append([vendor_name,regulated_vendor_name,matched_token])
        
       
        pd.DataFrame(matched_vendors,columns=[col_name,col_name,'Score']).to_csv('Str_Matching_'+ method.__name__ + '_on_' + col_name + '_' + date.today().strftime("%d%b%y") + '.csv', index=False)
        print('Excel has been created with the results!!')
        
    else:    
        print('Input the number in the range')
        sys.exit()    

else:
    print('Change the working directory!')
    sys.exit()



