# -*- coding: utf-8 -*-
"""
Created on Mon Mar 28 11:19:59 2022

@author: udyan.sachdev
"""

import pandas as pd
import numpy as np
import re
import sys
import scipy
import os
from datetime import date
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer
# pip install -U sentence-transformers

cwd = os.getcwd()
print("Current working directory: {0}".format(cwd))

inp = input('Do you want to continue - Y or N : ')
files_xls=[]

if (inp == 'Y' or inp == 'y' or inp =='Yes' or inp =='yes'):
    print(' ')
    file = os.listdir(cwd)
    files_xls1 = [f for f in file if (f[-3:] == 'csv') or (f[-3:] == 'xls')]
    files_xls2 = [f for f in file if f[-4:] == 'xlsx']

    files_xls = files_xls1 + files_xls2 
    
    if len(files_xls) == 0:
        print('No excel Found! Change the directory and try again')
        
    else:
        ###Taking input excel for data
        print('Here is the list of available excels : ')
        
        print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(files_xls))]))
        print(' ')
    
        num = int(input("Enter the serial number corresponding the required excel containing sentences : "))
        
        if files_xls[num][-3:] == 'csv':
            full_path = cwd + '\\' + files_xls[num]
    
            data = pd.read_csv(full_path)
            
        elif files_xls[num][-3:] == 'xls':
            full_path = cwd + '\\' + files_xls[num]
    
            data = pd.read_excel(full_path)
            
        else:
            full_path = cwd + '\\' + files_xls[num]
            
            wb = load_workbook(filename = files_xls[num])
          
            print(' ')
            print('Here is the list of available worksheet : ')
            print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(wb.sheetnames))]))
            print(' ')
            try:   
                sn = int(input("Enter the serial number corresponding the required Sheet name: "))
                
                data = pd.read_excel(full_path,sheet_name = wb.sheetnames[sn])
                
            except IndexError:
                print('Input the number in the range')
                sys.exit()
               
        
        print('')
        print('All the columns in this worksheet are : ')
        print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data.columns))]))
    
    
        rep1 = int(input('Enter the serial number of required column for semantic search : '))
        
        sentences = data[data.columns[rep1]].values.tolist()

        ###Taking input excel for query
        rep2 = int(input('Enter the serial number of containing queries/keywords : '))
        
        queries = data[data.columns[rep2]].values.tolist()
        
        for k in range(len(queries)):
            queries[k] = queries[k].lower()


        ###user input for semantic search type
        print('')
        print('')
        print('1 - For symmetric semantic search your query and the entries in your corpus are of about the same length and have the same amount of content. An example would be searching for similar questions: Your query could for example be “How to learn Python online?” and you want to find an entry like “How to learn Python on the web?”. For symmetric tasks, you could potentially flip the query and the entries in your corpus.')
        print('')
        print ('2 - For asymmetric semantic search, you usually have a short query (like a question or some keywords) and you want to find a longer paragraph answering the query. An example would be a query like “What is Python” and you wand to find the paragraph “Python is an interpreted, high-level and general-purpose programming language. Python’s design philosophy …”. For asymmetric tasks, flipping the query and the entries in your corpus usually does not make sense.')
        print('')
        print('Do you want to perform \n1 - Symmetric Semantic Search \n2 - Asymmetric Semantic Search')
        
        inp = int(input('1 or 2 : '))
        
        if inp == 1:
            print('')
            print('#### Model is loading ####')
            model = SentenceTransformer('all-mpnet-base-v2')
        
        else:
            print('')
            print('##### Model is loading #####')
            model = SentenceTransformer('msmarco-distilbert-base-tas-b')
        
        ###user input for top n sentences
        top_n = int(input('Enter the required closest n sentences of the corpus for each query : '))
        # Find the closest n sentences of the corpus for each query sentence based on cosine similarity
        number_top_matches = top_n #@param {type: "number"}
        
        input_data = []
        sentence = []
        score = []
        queryy = []
        
        for i in range(len(sentences)):
            clean_text = sentences[i].split(".")
            clean_text = list(filter(None, clean_text))
            df1 = pd.DataFrame(clean_text,columns = ['sentence'])
            clean_text = df1['sentence'].tolist()
            
            sentence_embeddings = model.encode(clean_text)
   
            query_embeddings = model.encode([queries[i]])
    
            for query, query_embedding in zip(queries[i], query_embeddings):
                distances = scipy.spatial.distance.cdist([query_embedding], sentence_embeddings, "cosine")[0]
            
                results = zip(range(len(distances)), distances)
                results = sorted(results, key=lambda x: x[1])
             
            for idx, distance in results[0:number_top_matches]:
                print(clean_text[idx].strip(), "(Cosine Score: %.4f)" % (1-distance))
                sentence.append(clean_text[idx].strip())
                score.append((1-distance))
                queryy.append(queries[i])                   
                input_data.append(data.iloc[i])      
                        
        output = pd.DataFrame(np.column_stack([sentence, score, queryy]), 
                                   columns=['nearest_sentence', 'score', 'query'])      
        output = output.join(pd.DataFrame(input_data).reset_index())
        output = output.drop(['index'], axis = 1)
        output.to_excel('Semantic_Search'+ '_' + date.today().strftime("%d%b%y") + '.xlsx', index=False)
                
else:
    sys.exit()