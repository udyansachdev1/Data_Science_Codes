# -*- coding: utf-8 -*-
"""
Created on Mon Nov  8 19:14:13 2021

@author: udyan.sachdev
"""


import pandas as pd
import os
import shutil
import numpy as np
import sys
from openpyxl import load_workbook
from tabulate import tabulate             #pip install tabulate

##Excel file should be present at the set cwd

cwd = os.getcwd()
print("Current working directory: {0}".format(cwd))

inp = input('Do you want to continue - Y or N : ')
files_xls=[]
if (inp == 'Y' or inp == 'y' or inp =='Yes' or inp =='yes'):
    print(' ')
    file = os.listdir(cwd)
    files_xls1 = [f for f in file if (f[-3:] == 'csv') or (f[-3:] == 'xls')]
    files_xls2 = [f for f in file if f[-4:] == 'xlsx']

    files_xls = files_xls1 + files_xls2 
    
    if len(files_xls) == 0:
        print('No excel Found! Change the directory and try again')
        
    else:
        try:
            print('Here is the list of available excels : ')

        
            print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(files_xls))]))
            print(' ')
        
            num = int(input("Enter the serial number corresponding the required excel: "))
            
            if files_xls[num][-3:] == 'csv':
                full_path = cwd + '\\' + files_xls[num]
        
                data = pd.read_csv(full_path)
                
            elif files_xls[num][-3:] == 'xls':
                full_path = cwd + '\\' + files_xls[num]
        
                data = pd.read_excel(full_path)
                
            else:
                full_path = cwd + '\\' + files_xls[num]
                
                wb = load_workbook(filename = files_xls[num])
              
                print(' ')
                print('Here is the list of available worksheet : ')
                print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(wb.sheetnames))]))
                print(' ')
                try:   
                    sn = int(input("Enter the serial number corresponding the required Sheet name: "))
                    
                    data = pd.read_excel(full_path,sheet_name = wb.sheetnames[sn])
                    
                except IndexError:
                    print('Input the number in the range')
                    sys.exit()   
                    
            folder = input('Enter the name of folder you want to store :')
            path = os.path.join(cwd, folder)
                    
        except IndexError:
            print('Input the number in the range')
            sys.exit()
    
            
    
        if os.path.exists(path):
            print(' ')
            print('The folder name is already present') 
            print(' ')
            fol = input('Do you want to create a new folder instead of replacing  -  Y or N :')
            if (fol == 'Y' or fol == 'y' or fol =='yes' or fol =='Yes'):
                    folder = input('Enter the name of folder you want to store :')
                    path = os.path.join(cwd, folder)
                    os.makedirs(path)
            else:
                print(' ')
                print('The existing folder is replaced')
                shutil.rmtree(path)
                os.makedirs(path)
        else:
            os.makedirs(path)  
        
        print('')
        print('All the columns in this worksheet are : ')
        print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data.columns))]))
    
    
        ###Removing bad characters
        rep1 = (input('Enter number corresponding the required column names from above in the required sequence of folders creation :'))
        rep1 = rep1.split(',')
        rep1 = list(map(int, rep1))
        bad_char = ['\<','\>','\:','\/','\"','\|','\?','\*','\\\\','\n']
        pattern = ('|'.join(bad_char))
        data = data.applymap(str)
        
        for i in rep1:
            x = data[data[data.columns[i]].str.contains(pattern, na=False)]
            print('')
            print('The column values having bad characters in' , data.columns[i], 'are : ', x[data.columns[i]].unique())
            print('')
                
    
        print(tabulate([['<',' '], ['>',' '],[':','-'],['/',' or '],['"',' '],['|',' _ '],['\\',''],['?','_'],['*','_'],['\n',' '],['nan','empty']], headers=['Original Bad Characters', 'Repaced with'], tablefmt='orgtbl'))
    
    
        rep = input('Do you want to continue with these changes - Y or N :')
        if (rep == 'Y' or rep == 'y' or rep =='Yes' or rep == 'yes'):
            
                for i in rep1:
                    data[data.columns[i]] = data[data.columns[i]].str.replace('<',' ',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace('>',' ',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace(':','-',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace('/',' or ',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace('"',' ',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace('\\',' ',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace('|',' _ ',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace('?','_',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace('*','_',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].str.replace('\n',' ',regex = True)
                    data[data.columns[i]] = data[data.columns[i]].replace(np.nan, 'empty', regex=True)      
                    data[data.columns[i]] = [a.strip(' ') for a in data[data.columns[i]]]   ##removing leading and trailing spaces
          
                
                store=[]
                for path_dir, group in data.groupby((data.columns[rep1]).tolist()):
                    store.append(path_dir)   
                    
                for k in range(len(store)):
                    p = os.path.join(path, *store[k])
                    if os.path.exists(p):
                         shutil.rmtree(p)
                    os.makedirs(p)
          
                
                print('')
                print('The folders has been created!')
                
        else:
            sys.exit()         
else:
    sys.exit()         