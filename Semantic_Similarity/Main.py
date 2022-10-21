# -*- coding: utf-8 -*-
"""
Created on Fri Jun  3 05:16:54 2022

@author: udyan.sachdev
"""

import pandas as pd
import numpy as np
import sys
import scipy
import os
from datetime import date
from openpyxl import load_workbook
from nltk.corpus import stopwords
# nltk.download('stopwords')
from gensim.utils import tokenize
from sentence_transformers import SentenceTransformer
# pip install -U sentence-transformers
        
###semantic search model       
model = SentenceTransformer('all-mpnet-base-v2')

cwd = os.getcwd()
print("Current working directory: {0}".format(cwd))

inp = input('Do you want to continue - Y or N : ')
files_xls=[]

if (inp == 'Y' or inp == 'y' or inp =='Yes' or inp =='yes'):
    print(' ')
    file = os.listdir(cwd)
    files_xls1 = [f for f in file if (f[-3:] == 'csv') or (f[-3:] == 'xls')]
    files_xls2 = [f for f in file if f[-4:] == 'xlsx']

    files_xls = files_xls1 + files_xls2 
    
    if len(files_xls) == 0:
        print('No excel Found! Change the directory and try again')
        
    else:
        ###Taking input excel for data
        print('Here is the list of available excels : ')
        
        print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(files_xls))]))
        print(' ')
    
        num = int(input("Enter the serial number corresponding the required excel containing sentences : "))
        
        if files_xls[num][-3:] == 'csv':
            full_path = cwd + '\\' + files_xls[num]
    
            data = pd.read_csv(full_path)
            
        elif files_xls[num][-3:] == 'xls':
            full_path = cwd + '\\' + files_xls[num]
    
            data = pd.read_excel(full_path)
            
        else:
            full_path = cwd + '\\' + files_xls[num]
            
            wb = load_workbook(filename = files_xls[num])
          
            print(' ')
            print('Here is the list of available worksheet : ')
            print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(wb.sheetnames))]))
            print(' ')
            try:   
                sn = int(input("Enter the serial number corresponding the required Sheet name: "))
                
                data = pd.read_excel(full_path,sheet_name = wb.sheetnames[sn])
                
            except IndexError:
                print('Input the number in the range')
                sys.exit()
               
        
        print('')
        print('All the columns in this worksheet are : ')
        print('\n'.join(['{}-{}'.format(i, val) for i, val in (enumerate(data.columns))]))
    
    
        rep1 = int(input('Enter the serial number of required column for semantic search : '))
        
        sentences = data[data.columns[rep1]].values.tolist()

        rep2 = int(input('Enter the serial number of containing queries/keywords : '))
        
        queries = data[data.columns[rep2]].values.tolist()
        
        for k in range(len(queries)):
            queries[k] = queries[k].lower()

        
        ###user input for top n sentences
        top_n = int(input('Enter the required closest n words of the corpus for each query : '))
        # Find the closest n sentences of the corpus for each query sentence based on cosine similarity
        number_top_matches = top_n #@param {type: "number"}
        
        input_data = []
        sentence = []
        score = []
        queryy = []

        for i in range(len(sentences)):
            
            clean_text = (' '.join(dict.fromkeys(sentences[i].split())))
            clean_text = list(tokenize(clean_text))
            clean_text = [word for word in clean_text if not word in stopwords.words()]
            
            sentence_embeddings = model.encode(clean_text)
            query_embeddings = model.encode([queries[i]])

            for query, query_embedding in zip(queries[i], query_embeddings):
                distances = scipy.spatial.distance.cdist([query_embedding], sentence_embeddings, "cosine")[0]
                results = zip(range(len(distances)), distances)
                results = sorted(results, key=lambda x: x[1])
             
            for idx, distance in results[0:number_top_matches]:
                print(clean_text[idx].strip(), "(Cosine Score: %.4f)" % (1-distance))
                sentence.append(clean_text[idx].strip())
                score.append((1-distance))
                queryy.append(queries[i])                   
                input_data.append(data.iloc[i])      
                            
        output = pd.DataFrame(np.column_stack([sentence, score, queryy]), 
                                   columns=['nearest_words', 'score', 'query'])      
        output = output.join(pd.DataFrame(input_data).reset_index())
        output = output.drop(['index'], axis = 1)
        output.to_excel('Semantic_Similarity'+ '_' + date.today().strftime("%d%b%y") + '.xlsx', index=False)

                
else:
    sys.exit()




